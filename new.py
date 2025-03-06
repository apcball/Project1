import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_template(ws):
    """จัดรูปแบบ template"""
    # สร้าง styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # จัดรูปแบบหัวคอลัมน์
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = border
    
    # กำหนดความกว้างคอลัมน์
    column_widths = {
        'A': 40,  # product_category_import
        'B': 40,  # parent_category
        'C': 20,  # costing_method
        'D': 15,  # income_account
        'E': 15   # expense_account
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

def process_category_path(display_name):
    """แยกชื่อ Category และ Parent Path จาก Display Name"""
    if pd.isna(display_name):
        return '', ''
    
    parts = str(display_name).strip().split('/')
    parts = [p.strip() for p in parts if p.strip()]  # ลบช่องว่างและรายการว่าง
    
    if not parts:
        return '', ''
    
    # ชื่อ category คือส่วนสุดท้าย
    category_name = parts[-1]
    
    # parent path คือทุกส่วนที่เหลือ รวมกันด้วย ' / '
    parent_path = ' / '.join(parts[:-1]) if len(parts) > 1 else ''
    
    return category_name, parent_path

def convert_data_to_template():
    # อ่านข้อมูลจากไฟล์ต้นฉบับ
    source_file = 'Data_file/Product Category update26022025.xlsx'
    df = pd.read_excel(source_file)
    print(f"Read {len(df)} rows from source file")
    
    # แปลงข้อมูล
    processed_data = []
    for _, row in df.iterrows():
        display_name = row['Display Name']
        category_name, parent_path = process_category_path(display_name)
        
        if category_name:  # ข้ามรายการที่ไม่มีชื่อ category
            processed_data.append({
                'product_category_import': category_name,
                'parent_category': parent_path,
                'costing_method': row['CostingMethod'] if pd.notna(row['CostingMethod']) else '',
                'income_account': row['Income Account'] if pd.notna(row['Income Account']) else '',
                'expense_account': row['Expense Account'] if pd.notna(row['Expense Account']) else ''
            })
    
    # สร้าง DataFrame ใหม่
    template_data = pd.DataFrame(processed_data)
    
    # บันทึกไฟล์ในโฟลเดอร์ Data_file
    output_file = 'Data_file/product_category_import.xlsx'
    
    # บันทึกข้อมูลลงในไฟล์ใหม่
    template_data.to_excel(output_file, sheet_name='Sheet1', index=False)
    
    # โหลดไฟล์เพื่อจัดรูปแบบ
    wb = load_workbook(output_file)
    ws = wb.active
    
    # จัดรูปแบบ template
    format_template(ws)
    
    # บันทึกไฟล์
    wb.save(output_file)
    
    print(f"\nData has been converted and saved to: {output_file}")
    print(f"Total rows processed: {len(template_data)}")
    print("\nColumns in the new file:")
    print("- product_category_import")
    print("- parent_category")
    print("- costing_method")
    print("- income_account")
    print("- expense_account")

if __name__ == "__main__":
    convert_data_to_template()