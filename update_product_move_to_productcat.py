"""Update product categories in Odoo by default_code -> product_category path.

This script was based on an earlier config file. It now reads a spreadsheet
or CSV containing columns `default_code` and `product_category` and updates the
corresponding product.template.categ_id in Odoo (via XML-RPC).

Defaults can be edited below or overridden via command-line arguments.
"""

# ==== Odoo Connection (defaults; can be overridden with CLI args) ====
URL = 'http://mogth.work:8069'
DB = 'MOG_LIVE'
USERNAME = 'apichart@mogen.co.th'
PASSWORD = '471109538'

# ==== I/O defaults ====
INPUT_PATH = 'Data_file/update_product_move_productcat.xlsx'   # expected columns: default_code, product_category
LOG_DIR = 'Data_file/logs'

import sys
import os
import argparse
import csv
import unicodedata
from xmlrpc import client


def normalize_name(s):
	if s is None:
		return ''
	s = str(s)
	s = unicodedata.normalize('NFKC', s)
	s = s.replace('\u00A0', ' ')
	return ' '.join(s.split()).strip()


def load_mapping(map_path):
	"""Load a CSV mapping file with columns: product_category,categ_id
	Returns dict mapping normalized product_category -> int(categ_id)
	"""
	if not map_path:
		return {}
	m = {}
	_, ext = os.path.splitext(map_path.lower())
	if ext in ('.csv', '.txt'):
		with open(map_path, newline='', encoding='utf-8-sig') as f:
			r = csv.DictReader(f)
			for row in r:
				key = normalize_name(row.get('product_category') or row.get('product category') or '')
				val = row.get('categ_id') or row.get('categid') or row.get('categ')
				if not key or not val:
					continue
				try:
					m[key] = int(val)
				except Exception:
					continue
	else:
		raise RuntimeError('Mapping file must be CSV')
	return m


def read_rows(path):
	"""Yield dict rows with keys 'default_code' and 'product_category'. Supports CSV and XLSX.
	For XLSX, openpyxl is used. If not installed, instruct the user.
	"""
	_, ext = os.path.splitext(path.lower())
	if ext in ('.csv', '.txt'):
		with open(path, newline='', encoding='utf-8-sig') as f:
			reader = csv.DictReader(f)
			for r in reader:
				yield r
		return

	if ext in ('.xls', '.xlsx'):
		try:
			from openpyxl import load_workbook
		except Exception:
			raise RuntimeError('openpyxl is required to read XLSX files. Install with: pip install openpyxl')
		wb = load_workbook(filename=path, read_only=True)
		ws = wb.active
		# get header
		rows = ws.iter_rows(values_only=True)
		try:
			header = [h if h is not None else '' for h in next(rows)]
		except StopIteration:
			return
		# normalize header
		header = [str(h).strip() for h in header]
		for row in rows:
			row = [c if c is not None else '' for c in row]
			d = dict(zip(header, row))
			yield d
		return

	raise RuntimeError('Unsupported file extension: ' + ext)


def connect_xmlrpc(url, db, username, password):
	if not url.endswith('/'):
		url = url + '/'
	common = client.ServerProxy(url + 'xmlrpc/2/common')
	uid = common.authenticate(db, username, password, {})
	if not uid:
		raise RuntimeError('Authentication failed; check credentials')
	models = client.ServerProxy(url + 'xmlrpc/2/object')
	return uid, models


def find_or_create_category(models, db, uid, password, path):
	parts = [p.strip() for p in str(path).split('/') if p.strip()]
	import unicodedata

	def normalize_name(s):
		if s is None:
			return ''
		# normalize unicode, replace non-breaking spaces, collapse whitespace
		s = str(s)
		s = unicodedata.normalize('NFKC', s)
		s = s.replace('\u00A0', ' ')
		s = ' '.join(s.split())
		return s.strip()

	parts = [normalize_name(p) for p in str(path).split('/') if normalize_name(p)]
	parent_id = False
	last_id = None

	# If path starts with a common top-level token like 'All', drop it for matching convenience
	if parts and parts[0].lower() == 'all':
		parts = parts[1:]

	# Try matching full joined path as a single category name (handles cases where names store the full path)
	if parts:
		full_join = ' / '.join(parts)
		try:
			res_full = models.execute_kw(db, uid, password, 'product.category', 'search', [[('name', 'ilike', full_join)]], {'limit': 5})
			if res_full:
				cands = models.execute_kw(db, uid, password, 'product.category', 'read', [res_full], {'fields': ['id', 'name']})
				# pick the candidate whose name best contains the full_join
				for c in cands:
					if full_join in normalize_name(c.get('name')):
						print(f'Found category by full-path match: "{full_join}" -> id={c.get("id")}')
						return c.get('id')
				# fallback to first candidate
				print(f'Using candidate for full-path ilike match: "{full_join}" -> id={cands[0].get("id")}')
				return cands[0].get('id')
		except Exception:
			pass

	for name in parts:
		# Search for category with this name and parent.
		# Try exact match first, then a case-insensitive 'ilike' fallback.
		domain_exact = [[('name', '=', name), ('parent_id', '=', parent_id)]]
		res = models.execute_kw(db, uid, password, 'product.category', 'search', domain_exact, {'limit': 1})
		if res:
			cat_id = res[0]
		else:
			# fallback: ilike (case-insensitive, may match substrings). Limit to parent scope.
			domain_ilike = [[('name', 'ilike', name), ('parent_id', '=', parent_id)]]
			res_ilike = models.execute_kw(db, uid, password, 'product.category', 'search', domain_ilike, {'limit': 5})
			if res_ilike:
				# try to find the best candidate with normalized exact equality
				candidates = models.execute_kw(db, uid, password, 'product.category', 'read', [res_ilike], {'fields': ['id', 'name', 'parent_id']})
				chosen = None
				# If any candidate contains the remaining path as a substring, use it as final
				full_remaining = ' / '.join(parts[parts.index(name):])
				for c in candidates:
					cname_norm = normalize_name(c.get('name'))
					if full_remaining and full_remaining in cname_norm:
						print(f'Using candidate that contains full remaining path "{full_remaining}" -> id={c.get("id")}')
						chosen = c['id']
						break
				for c in candidates:
					if normalize_name(c.get('name')) == name:
						chosen = c['id']
						break
				if not chosen:
					# no exact normalized match; pick first but warn
					chosen = candidates[0]['id']
					print(f'Warning: ambiguous category name "{name}" under parent={parent_id}; candidates: {[c.get("name") for c in candidates]}; using id={chosen}')
				else:
					print(f'Found category by ilike and normalized match: "{name}" -> id={chosen}')
				cat_id = chosen
			else:
				# Do not create categories automatically. Skip if not found.
				print(f'Category level "{name}" not found (parent={parent_id}); skipping (auto-creation disabled)')
				return None

			# If still not found, try global candidate search by matching the last segment against any category
			if not res and not res_ilike:
				# Only attempt this fallback when matching deeper levels (not first level) to avoid spurious matches
				if parts:
					last_segment = name
					cand_domain = [[('name', 'ilike', last_segment)]]
					cand_ids = models.execute_kw(db, uid, password, 'product.category', 'search', cand_domain, {'limit': 50})
					if cand_ids:
						# For each candidate, walk up parents and compare against reversed parts
						for cid in cand_ids:
							anc = []
							cur = cid
							while cur:
								r = models.execute_kw(db, uid, password, 'product.category', 'read', [[cur], ['id', 'name', 'parent_id']])
								if not r:
									break
								rec = r[0]
								anc.append(normalize_name(rec.get('name')))
								p = rec.get('parent_id')
								if not p:
									break
								if isinstance(p, list):
									cur = p[0]
								else:
									cur = p
							# Now anc is [child, parent, grandparent,...]
							rev_parts = list(reversed(parts[:parts.index(name)+1]))
							# Compare the sequence; allow match if anc startswith rev_parts
							if len(anc) >= len(rev_parts) and anc[:len(rev_parts)] == rev_parts:
								cat_id = cid
								print(f'Found category by ancestor chain matching: "{name}" -> id={cid}')
								break
						if cat_id:
							# found candidate
							pass

		last_id = cat_id
		parent_id = cat_id
	return last_id


def update_product_category(models, db, uid, password, default_code, category_id, dry_run=False):
	domain = [[('default_code', '=', default_code)]]
	prod_ids = models.execute_kw(db, uid, password, 'product.product', 'search', domain)
	if not prod_ids:
		print(f'No product found with default_code="{default_code}"')
		return False
	if len(prod_ids) > 1:
		print(f'Warning: {len(prod_ids)} products found for default_code="{default_code}"; using first id={prod_ids[0]}')
	prod_id = prod_ids[0]

	read = models.execute_kw(db, uid, password, 'product.product', 'read', [[prod_id], ['product_tmpl_id']])
	if not read:
		print(f'Failed to read product.product id={prod_id}')
		return False
	tmpl_field = read[0].get('product_tmpl_id')
	if isinstance(tmpl_field, list):
		tmpl_id = tmpl_field[0]
	else:
		tmpl_id = tmpl_field

	print(f'Updating product {default_code}: product.product id={prod_id} -> template id={tmpl_id} set categ_id={category_id}')
	if dry_run:
		return True
	try:
		models.execute_kw(db, uid, password, 'product.template', 'write', [[tmpl_id], {'categ_id': category_id}])
	except Exception as e:
		print(f'Error writing template id={tmpl_id}: {e}')
		return False
	return True


def process_file(path, url, db, username, password, map_path=None, dry_run=False):
	uid, models = connect_xmlrpc(url, db, username, password)
	print(f'Connected to {url} db={db} uid={uid}')
	updated = 0
	skipped = 0
	mapping = load_mapping(map_path)
	for row in read_rows(path):
		# try multiple header variants
		default_code = (row.get('default_code') or row.get('Number') or row.get('number') or row.get('defaultcode') or '').strip()
		category_path = (row.get('product_category') or row.get('product category') or row.get('expense_account') or row.get('product_category_path') or '').strip()
		if not default_code:
			print('Skipping row with empty default_code')
			skipped += 1
			continue
		if not category_path:
			print(f'Skipping {default_code}: empty product_category')
			skipped += 1
			continue
		# try mapping first
		mapped = mapping.get(normalize_name(category_path))
		if mapped:
			cat_id = mapped
			print(f'Category path matched via mapping -> id={cat_id}')
		else:
			cat_id = find_or_create_category(models, db, uid, password, category_path)
		if not cat_id:
			skipped += 1
			continue
		ok = update_product_category(models, db, uid, password, default_code, cat_id, dry_run=dry_run)
		if ok:
			updated += 1
		else:
			skipped += 1
	print(f'Done. Updated: {updated}, Skipped: {skipped}')


def main():
	parser = argparse.ArgumentParser(description='Update product categories by default_code -> category path')
	parser.add_argument('--url', default=URL)
	parser.add_argument('--db', default=DB)
	parser.add_argument('--username', default=USERNAME)
	parser.add_argument('--password', default=PASSWORD)
	parser.add_argument('--input', default=INPUT_PATH, help='CSV or XLSX file path')
	parser.add_argument('--map', dest='map', default=None, help='Optional CSV mapping file: product_category,categ_id')
	parser.add_argument('--dry-run', action='store_true')
	args = parser.parse_args()
	try:
		process_file(args.input, args.url, args.db, args.username, args.password, map_path=args.map, dry_run=args.dry_run)
	except Exception as e:
		print('Error:', e)
		sys.exit(1)


if __name__ == '__main__':
	main()